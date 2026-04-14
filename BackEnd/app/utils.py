from datetime import date, timedelta
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font,Border, Side
from openpyxl.utils import get_column_letter

def semana_actual():
    hoy = date.today()
    iso = hoy.isocalendar()
    num_semana = iso[1]
    año = iso[0]  # también usar el año ISO, no hoy.year

    # Anclar al lunes de la semana ISO
    lunes = hoy - timedelta(days=hoy.weekday())
    mes = lunes.month

    return f"{año}_{mes:02d}_semana_{num_semana}", año, num_semana

def procesarDatosTecnico(datos: dict, nombre: str, trabajo:str):
    dicDataModificar={
        "NOMBRE":"",
        "JOB":"",
        "JOB_NAME": "",
        "PORCENTAJE_TECNICO": 0,
        "VALOR_SERVICIO":0,
        "MINIMO":0,
        "TIPO_PAGO":["CASH", "CC", "MIXTO"],
        "VALOR_TARJETA":0,
        "VALOR_CASH":0,
        "PORCENTAJE_CC":0,
        "PARTES_GIL": 0,
        "PARTES_TECNICO": 0,
        "TECH":0,
        "SUBTOTAL":0,
        "TOTAL":0,
        "NOTAS": []
    }

    #Nombre del tenico
    dicDataModificar["NOMBRE"] = nombre

    #PORCENTAJE DEL TENICO

    if(datos["porcentaje_adicional_empresa"] != 0):
        dicDataModificar["NOTAS"].append(
            f"Este tecnico tiene un porcentaje adicional para ciertos servicios de {datos['porcentaje_adicional_empresa']}%"
        )

    dicDataModificar["PORCENTAJE_TECNICO"] = datos["porcentaje_tenico"]

    #PORCENTAJE CC

    dicDataModificar["PORCENTAJE_CC"] = datos["porcentaje_cc"]

    #MINIMO

    dicDataModificar["MINIMO"] = datos["minimo"]
    

    
def obtener_rango_semana():
    hoy = date.today()

    # lunes de esta semana
    print(hoy)
    fecha_inicio = hoy - timedelta(days=hoy.weekday())
    print(timedelta(days=hoy.weekday()))

    # domingo de esta semana
    fecha_fin = fecha_inicio + timedelta(days=6)

    return fecha_inicio, fecha_fin

def construcionTablaResultado(df):
    df.columns = df.columns.str.upper()
    dfResultados = {
        "NOMBRE": [
            "TOTAL JOBS",
            "TOTAL CASH",
            "TOTAL CC",
            "TOTAL SALES",
            "TOTAL PARTS",
            "AVERAGE SALES",
        ],
        "RESULTADO": []
    }

    # TOTAL JOBS
    totalJobs = len(df)
    dfResultados["RESULTADO"].append(totalJobs)

    # CASH
    dfCash = df[df["TIPO PAGO"] == "CASH"]
    sumaCash = dfCash["VALOR SERVICIO"].sum()

    # CC
    dfCC = df[df["TIPO PAGO"] == "CC"]
    sumaCC = dfCC["VALOR SERVICIO"].sum()

    # MIXTO
    dfMixto = df[df["TIPO PAGO"] == "MIXTO"]

    sumaCash += dfMixto["VALOR EFECTIVO"].sum()
    sumaCC += dfMixto["VALOR TARJETA"].sum()

    dfResultados["RESULTADO"].append(sumaCash)
    dfResultados["RESULTADO"].append(sumaCC)

    # TOTAL SALES
    totalSales = sumaCash + sumaCC
    dfResultados["RESULTADO"].append(totalSales)

    # TOTAL PARTS
    totalPartesGil = df["PARTES GIL"].sum()
    totalPartesTecnico = df["PARTES TECNICO"].sum()

    totalParts = totalPartesGil + totalPartesTecnico
    dfResultados["RESULTADO"].append(totalParts)

    # AVERAGE SALES
    averageSales = totalSales / totalJobs if totalJobs else 0
    dfResultados["RESULTADO"].append(averageSales)

    return pd.DataFrame(dfResultados)




    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    # =========================
    # TABLA 1 (REGISTROS)
    # =========================

    ultima_columna = get_column_letter(len(df.columns))
    ultima_fila = len(df) + 1

    rango_tabla1 = f"A1:{ultima_columna}{ultima_fila}"

    tabla1 = Table(displayName="TablaRegistros", ref=rango_tabla1)

    estilo_tabla1 = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    tabla1.tableStyleInfo = estilo_tabla1
    ws.add_table(tabla1)

    # =========================
    # FILAS ROJAS SI TOTAL < 0
    # =========================

    col_total = df.columns.get_loc("TOTAL") + 1

    rojo_claro = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in range(2, ultima_fila + 1):

        valor_total = ws.cell(row=row, column=col_total).value

        if valor_total is not None and valor_total < 0:

            for col in range(1, len(df.columns) + 1):
                ws.cell(row=row, column=col).fill = rojo_claro

    # =========================
    # TABLA 2 (RESULTADOS)
    # =========================

    inicio_tabla2 = fila_inicio + 1
    fin_tabla2 = inicio_tabla2 + len(dfResultado)

    rango_tabla2 = f"A{inicio_tabla2}:B{fin_tabla2}"

    tabla2 = Table(displayName="TablaResultados", ref=rango_tabla2)

    estilo_tabla2 = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    tabla2.tableStyleInfo = estilo_tabla2
    ws.add_table(tabla2)

    nuevo_output = BytesIO()
    wb.save(nuevo_output)
    nuevo_output.seek(0)

    return nuevo_output

def estilizar_excel(output, df, dfResultado, fila_inicio):

    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    # =========================
    # ANCHOS DE COLUMNA
    # =========================
    anchos = {
        "A": 12,  # NAME
        "B": 14,  # ID JOB
        "C": 6,   # %
        "D": 14,  # PAYMENT TYPE
        "E": 10,  # SALES
        "F": 8,   # 4%CC
        "G": 10,  # GIL PARTS
        "H": 12,  # TECH PARTS
        "I": 8,   # TECH
        "J": 12,  # TOTAL
    }
    for col_letra, ancho in anchos.items():
        ws.column_dimensions[col_letra].width = ancho

    # =========================
    # FORMATO MONEDA $#,##0.00 en columnas numéricas
    # =========================
    fmt_moneda = '$#,##0.00_);[Red]($#,##0.00);$ -'
    fmt_pct    = '0"%"'

    col_indices = {col: idx + 1 for idx, col in enumerate(df.columns)}

    cols_moneda = ["SALES", "4%CC", "GIL PARTS", "TECH PARTS", "TECH", "TOTAL"]
    if "CASH" in df.columns:
        cols_moneda += ["CASH", "CC"]

    ultima_fila_datos = len(df) + 1

    for col_name in cols_moneda:
        if col_name in col_indices:
            col_num = col_indices[col_name]
            for row in range(2, ultima_fila_datos + 1):
                ws.cell(row=row, column=col_num).number_format = fmt_moneda

    # Columna % como porcentaje sin símbolo (ej: 30)
    if "%" in col_indices:
        col_pct = col_indices["%"]
        for row in range(2, ultima_fila_datos + 1):
            ws.cell(row=row, column=col_pct).number_format = fmt_pct

    # =========================
    # TABLA 1 (REGISTROS)
    # =========================
    ultima_columna = get_column_letter(len(df.columns))
    rango_tabla1 = f"A1:{ultima_columna}{ultima_fila_datos}"
    tabla1 = Table(displayName="TablaRegistros", ref=rango_tabla1)
    tabla1.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabla1)

    # =========================
    # FILAS ROJAS SI TOTAL < 0
    # =========================
    col_total = col_indices.get("TOTAL", len(df.columns))
    rojo_claro = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in range(2, ultima_fila_datos + 1):
        valor_total = ws.cell(row=row, column=col_total).value
        if valor_total is not None and valor_total < 0:
            for col in range(1, len(df.columns) + 1):
                ws.cell(row=row, column=col).fill = rojo_claro

    # =========================
    # TABLA 2 (RESULTADOS)
    # =========================
    inicio_tabla2 = fila_inicio + 1
    fin_tabla2    = inicio_tabla2 + len(dfResultado)
    rango_tabla2  = f"A{inicio_tabla2}:B{fin_tabla2}"
    tabla2 = Table(displayName="TablaResultados", ref=rango_tabla2)
    tabla2.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabla2)

    # =========================
    # COLORES FILAS RESULTADOS  (exactos de la imagen)
    # =========================
    colores_resultados = [
        ("FFFF00", "000000"),  # TOTAL JOBS    → amarillo, texto negro
        ("00FF00", "000000"),  # TOTAL CASH    → verde, texto negro
        ("FF0000", "FFFFFF"),  # TOTAL CC      → rojo, texto blanco
        ("008B8B", "FFFFFF"),  # TOTAL SALES   → cyan oscuro, texto blanco
        ("6A5ACD", "FFFFFF"),  # TOTAL PARTS   → azul morado, texto blanco
        ("FF8C00", "FFFFFF"),  # AVERAGE SALES → naranja oscuro, texto blanco
    ]

    for i, (bg, fg) in enumerate(colores_resultados):
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        font = Font(bold=True, color=fg)
        fila_excel = inicio_tabla2 + 1 + i
        for col in range(1, 3):
            cell = ws.cell(row=fila_excel, column=col)
            cell.fill = fill
            cell.font = font

    # Formato moneda en columna B de resultados (excepto TOTAL JOBS y AVERAGE)
    fmt_resultado = '$#,##0.00_);[Red]($#,##0.00);$ -'
    for i in range(len(dfResultado)):
        ws.cell(row=inicio_tabla2 + 1 + i, column=2).number_format = fmt_resultado

    # =========================
    # BALANCED TECH (verde oscuro, texto blanco negrita)
    # =========================
    col_balance  = 9
    fila_balance = inicio_tabla2

    verde_oscuro = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    fuente_blanca_bold = Font(bold=True, color="FFFFFF")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in range(fila_balance, fila_balance + 2):
        for col in range(col_balance, col_balance + 2):
            cell = ws.cell(row=row, column=col)
            cell.border    = thin_border
            cell.fill      = verde_oscuro
            cell.font      = fuente_blanca_bold

    # Formato moneda al valor de BALANCED TECH
    ws.cell(row=fila_balance + 1, column=col_balance + 1).number_format = fmt_moneda

    # =========================
    # ALINEACIÓN GENERAL
    # =========================
    center = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=1, max_row=fin_tabla2 + 2):
        for cell in row:
            cell.alignment = center

    nuevo_output = BytesIO()
    wb.save(nuevo_output)
    nuevo_output.seek(0)
    return nuevo_output