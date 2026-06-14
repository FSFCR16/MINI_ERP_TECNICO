"""
Generación de Excel del módulo VIRGINIA. Independiente de utils.py.

Formato (basado en la plantilla de cierre "AMAN ATLANTA"):
  Columnas: Job | % | SALES | parts | tecH | SUBTOTAL | TOTAL | NOTES
  - SALES con signo: CASH (+), CC (-), MIXTO (+ total)
  - % como fracción: CASH → % técnico ; CC → % gil (100-téc) ; MIXTO → 0
  - SUBTOTAL = SALES (mismo valor con signo)
  - TOTAL = el total ya calculado por Virginia (respeta mínimo / %propio / MIXTO)
  - NOTES = método de pago ; en MIXTO: "<efectivo> CASH / <tarjeta> CC"
  - parts / tecH: columnas presentes (vacías en filas normales) para respetar el layout
  - Valores en SALES/SUBTOTAL/TOTAL: verde si >= 0, rojo si < 0
  - Fila PARTS (cyan) al final de la tabla (solo aporta al BALANCED TECH)
  Bloque de totales (desglose DINÁMICO por método):
    TOTAL JOBS · TOTAL CASH · TOTAL <cada método CC presente> · TOTAL SALES · AVERAGE SALES
  BALANCED TECH = Σ(registros.total) + Σ(partes.valor × % / 100)
"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ── formatos ──
FMT_MONEDA = '"$"#,##0.00'
# verde si positivo, rojo si negativo, negro en cero
FMT_MONEDA_COLOR = '[Green]"$"#,##0.00;[Red]-"$"#,##0.00;"$"#,##0.00'
FMT_PCT = '0.00%'
FMT_ENTERO = '#,##0'

# ── colores (fill) ──
C_HEADER = "FCE4D6"   # peach
C_PARTS = "00FFFF"    # cyan
C_JOBS = "FFFF00"     # amarillo
C_CASH = "00FF00"     # verde
C_CC = "FF0000"       # rojo
C_SALES = "00FFFF"    # cyan
C_AVG = "FF9900"      # naranja
C_BAL = "00B050"      # verde balanced

COLUMNAS = ["Job", "%", "SALES", "parts", "tecH", "SUBTOTAL", "TOTAL", "NOTES"]
# índices 1-based
COL_JOB, COL_PCT, COL_SALES, COL_PARTS, COL_TECH, COL_SUBTOTAL, COL_TOTAL, COL_NOTES = 1, 2, 3, 4, 5, 6, 7, 8

_thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_center = Alignment(horizontal="center", vertical="center")


def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _fila_sales_pct_notes(r):
    """Devuelve (sales_con_signo, pct_fraccion, notes) según el tipo de pago."""
    tipo = (r.tipo_pago or "").upper()
    sales = r.valor_servicio or 0
    pct_tec = (r.porcentaje_tecnico or 0) / 100

    if tipo == "CC":
        return -sales, (100 - (r.porcentaje_tecnico or 0)) / 100, (r.metodo_pago or "CC")
    if tipo == "MIXTO":
        ef = r.valor_efectivo or 0
        ta = r.valor_tarjeta or 0
        return sales, 0, f"{ef:g} CASH / {ta:g} CC"
    # CASH (default)
    return sales, pct_tec, (r.metodo_pago or "CASH")


def _resumen_por_metodo(registros):
    """
    Devuelve (total_cash, cc_por_metodo: dict, total_sales).
    total_cash    = Σ ventas CASH + Σ efectivo de MIXTO
    cc_por_metodo = {nombre_metodo: Σ ventas CC}  (DINÁMICO según los métodos de la tabla;
                    incluye 'MIXTO' por la parte tarjeta)
    """
    total_cash = 0.0
    cc = {}
    for r in registros:
        tipo = (r.tipo_pago or "").upper()
        if tipo == "CASH":
            total_cash += r.valor_servicio or 0
        elif tipo == "CC":
            nombre = (r.metodo_pago or "CC").upper()
            cc[nombre] = cc.get(nombre, 0) + (r.valor_servicio or 0)
        elif tipo == "MIXTO":
            total_cash += r.valor_efectivo or 0
            if (r.valor_tarjeta or 0):
                cc["MIXTO"] = cc.get("MIXTO", 0) + (r.valor_tarjeta or 0)
    total_sales = total_cash + sum(cc.values())
    return total_cash, cc, total_sales


def generarExcelVirginia(registros, partes):
    """
    registros: lista de RegistroVirginia (ORM)
    partes:    lista de ParteVirginia (ORM)
    Devuelve BytesIO con el .xlsx listo.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cierre"

    ncols = len(COLUMNAS)

    # ── header ──
    for idx, nombre in enumerate(COLUMNAS, start=1):
        c = ws.cell(row=1, column=idx, value=nombre)
        c.fill = _fill(C_HEADER)
        c.font = Font(bold=True)
        c.border = _thin
        c.alignment = _center

    # ── filas de registros ──
    fila = 2
    for r in registros:
        sales, pct, notes = _fila_sales_pct_notes(r)
        ws.cell(row=fila, column=COL_JOB, value=r.job_name)
        ws.cell(row=fila, column=COL_PCT, value=pct).number_format = FMT_PCT
        ws.cell(row=fila, column=COL_SALES, value=sales).number_format = FMT_MONEDA_COLOR
        ws.cell(row=fila, column=COL_PARTS).number_format = FMT_MONEDA
        ws.cell(row=fila, column=COL_TECH).number_format = FMT_MONEDA
        ws.cell(row=fila, column=COL_SUBTOTAL, value=sales).number_format = FMT_MONEDA_COLOR
        ws.cell(row=fila, column=COL_TOTAL, value=r.total or 0).number_format = FMT_MONEDA_COLOR
        ws.cell(row=fila, column=COL_NOTES, value=notes)
        for col in range(1, ncols + 1):
            ws.cell(row=fila, column=col).border = _thin
            ws.cell(row=fila, column=col).alignment = _center
        fila += 1

    ultima_fila_datos = fila - 1

    # ── fila PARTS (cyan) ──
    balance_partes = sum((p.valor or 0) * ((p.porcentaje_tecnico or 0) / 100) for p in partes)
    valor_total_partes = sum(p.valor or 0 for p in partes)
    pcts = {p.porcentaje_tecnico for p in partes if p.porcentaje_tecnico is not None}
    pct_partes = (pcts.pop() / 100) if len(pcts) == 1 else None

    if valor_total_partes:
        fp = fila
        ws.cell(row=fp, column=COL_JOB, value="PARTS")
        if pct_partes is not None:
            ws.cell(row=fp, column=COL_PCT, value=pct_partes).number_format = FMT_PCT
        ws.cell(row=fp, column=COL_SALES, value=valor_total_partes).number_format = FMT_MONEDA_COLOR
        ws.cell(row=fp, column=COL_SUBTOTAL, value=valor_total_partes).number_format = FMT_MONEDA_COLOR
        ws.cell(row=fp, column=COL_TOTAL, value=balance_partes).number_format = FMT_MONEDA_COLOR
        for col in range(1, ncols + 1):
            cell = ws.cell(row=fp, column=col)
            cell.fill = _fill(C_PARTS)
            cell.border = _thin
            cell.alignment = _center
        fila += 1

    # ── bloque de totales (col A label / col B valor) — DINÁMICO por método ──
    total_cash, cc_por_metodo, total_sales = _resumen_por_metodo(registros)
    total_jobs = len(registros)
    average = (total_sales / total_jobs) if total_jobs else 0

    # filas: (label, valor, color_fill, font_color, formato)
    filas_tot = [("TOTAL JOBS", total_jobs, C_JOBS, "000000", FMT_ENTERO),
                 ("TOTAL CASH", total_cash, C_CASH, "000000", FMT_MONEDA)]
    for nombre, monto in cc_por_metodo.items():
        filas_tot.append((f"TOTAL {nombre}", monto, C_CC, "FFFFFF", FMT_MONEDA))
    filas_tot.append(("TOTAL SALES", total_sales, C_SALES, "000000", FMT_MONEDA))
    filas_tot.append(("AVERAGE SALES", average, C_AVG, "000000", FMT_MONEDA))

    inicio_tot = fila + 2
    for i, (label, monto, bg, fg, fmt) in enumerate(filas_tot):
        f = inicio_tot + i
        cl = ws.cell(row=f, column=1, value=label)
        cv = ws.cell(row=f, column=2, value=monto)
        cv.number_format = fmt
        for cell in (cl, cv):
            cell.fill = _fill(bg)
            cell.font = Font(bold=True, color=fg)
            cell.border = _thin
            cell.alignment = _center

    # ── BALANCED TECH (en la columna TOTAL, al lado del bloque) ──
    balanced = sum((r.total or 0) for r in registros) + balance_partes
    bl = ws.cell(row=inicio_tot, column=COL_TOTAL, value="BALANCED TECH")
    bv = ws.cell(row=inicio_tot + 1, column=COL_TOTAL, value=balanced)
    bv.number_format = FMT_MONEDA
    for cell in (bl, bv):
        cell.fill = _fill(C_BAL)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = _thin
        cell.alignment = _center

    # ── anchos ──
    anchos = {
        COL_JOB: 12, COL_PCT: 9, COL_SALES: 14, COL_PARTS: 11,
        COL_TECH: 11, COL_SUBTOTAL: 14, COL_TOTAL: 16, COL_NOTES: 24,
    }
    for col, w in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
